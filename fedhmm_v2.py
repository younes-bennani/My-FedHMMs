"""
fedhmm.py
═════════════════════════════════════════════════════════════════════
Federated Hidden Markov Models for Privacy-Preserving Analysis of
Governance Action Sequences.

Implements the four algorithms described in:
  Hemmati & Bennani, "Federated Hidden Markov Models for
  Privacy-Preserving Analysis of Governance Action Sequences", 2025.

Algorithms
----------
  [0] Centralised     : Baum-Welch on all pooled data (oracle upper bound)
  [1] Local-HMM       : independent Baum-Welch per user (privacy baseline)
  [2] FedAvg-HMM      : Algorithm 2 — weighted parameter averaging
  [3] FedEM-HMM       : Algorithm 1 — raw sufficient-statistics aggregation
  [4] Personalised    : Algorithm 3 — FedEM global + local fine-tuning + α-interpolation

Mathematics (matching the paper exactly)
-----------------------------------------
  Observation alphabet : X = {ADD=0, UPDATE=1, DELETE=2}
  HMM parameters       : θ = (π, A, B)
                         π ∈ ΔK-1, A ∈ [0,1]^{K×K}, B ∈ [0,1]^{K×|X|}
  Sufficient statistics:
    Γ_i   = Σ_s γ^(s)_1(i)
    Ξ_{ij} = Σ_s Σ_{t=1}^{T_s-1} ξ^(s)_t(i,j)     ← t up to T_s-1
    Φ_{io} = Σ_s Σ_{t=1}^{T_s}   γ^(s)_t(i) 1[o^(s)_t = o]
  M-step   : π̂ = Γ/‖Γ‖₁,  Â_{ij} = Ξ_{ij}/Σ_{j'}Ξ_{ij'},  B̂_{io} = Φ_{io}/Σ_{o'}Φ_{io'}
  FedEM    : Γ = Σ_i Γ^(i)  (simple sum, no normalisation)
  BIC      : -2 log L̂ + p(K)·ln(N_tot),  p(K) = (K-1) + K(K-1) + K(|X|-1)

Usage
-----
  python fedhmm.py sessions.jsonl [output.xlsx]

  Default output: fedhmm_results.xlsx
"""

import json
import sys
import math
import warnings
from pathlib import Path
from collections import defaultdict
from dataclasses import dataclass, field
from typing import List, Dict, Tuple, Optional

import numpy as np
from hmmlearn.hmm import CategoricalHMM
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────────────────────────────
# Constants
# ─────────────────────────────────────────────────────────────────────

ACTION_MAP: Dict[str, int] = {"ADD": 0, "UPDATE": 1, "DELETE": 2}
ACTION_SYM: Dict[int, str] = {0: "A", 1: "U", 2: "D"}
N_OBS: int  = 3          # |X|
EPS: float  = 1e-12      # numerical floor for log

# Hyperparameters (matching Section 7 of the paper)
DEFAULT_K:      int   = 4      # BIC-optimal number of states
K_RANGE               = range(2, 6)
N_ROUNDS:       int   = 20     # R — federation rounds
M_LOCAL:        int   = 5      # M — local BW iterations (FedAvg)
M_PERSONAL:     int   = 30     # M_p — personalisation fine-tuning
ALPHA_PERS:     float = 0.30   # α — personalisation coefficient
MIN_SESSIONS:   int   = 3      # minimum sessions per user
TEST_RATIO:     float = 0.25   # train/test split
N_INIT:         int   = 3      # random re-starts for model selection

# Formatting
BLUE_DARK = "2B4590"; BLUE_LIGHT = "D9E1F7"; ALT = "EEF2FA"
WHITE = "FFFFFF"; GREY = "F5F5F5"
THIN = Border(**{s: Side(style="thin", color="C0C8D8")
                 for s in ("left", "right", "top", "bottom")})

ALGO_COLORS = {
    "Centralised":   "9B59B6",
    "Local":         "E74C3C",
    "FedAvg":        "F39C12",
    "FedEM":         "27AE60",
    "Personalised":  "2980B9",
}


# ─────────────────────────────────────────────────────────────────────
# HMM parameter container
# ─────────────────────────────────────────────────────────────────────

@dataclass
class HMMParams:
    """Container for HMM parameters θ = (π, A, B)."""
    startprob: np.ndarray   # shape (K,)
    transmat:  np.ndarray   # shape (K, K)  — row-stochastic
    emitprob:  np.ndarray   # shape (K, |X|) — row-stochastic
    K: int

    @staticmethod
    def random_init(K: int, seed: int = 0) -> "HMMParams":
        rng = np.random.default_rng(seed)
        sp = rng.dirichlet(np.ones(K))
        tr = rng.dirichlet(np.ones(K), size=K)
        ep = rng.dirichlet(np.ones(N_OBS), size=K)
        return HMMParams(sp, tr, ep, K)

    def to_log(self):
        return (np.log(np.clip(self.startprob, EPS, 1)),
                np.log(np.clip(self.transmat,  EPS, 1)),
                np.log(np.clip(self.emitprob,  EPS, 1)))

    def copy(self) -> "HMMParams":
        return HMMParams(self.startprob.copy(),
                         self.transmat.copy(),
                         self.emitprob.copy(),
                         self.K)

    def interpolate(self, other: "HMMParams", alpha: float) -> "HMMParams":
        """(1-α)·self + α·other, component-wise.
        A convex combination of row-stochastic matrices is already row-stochastic;
        the explicit renormalisation below is kept only for floating-point safety.
        (Paper v3, Alg. 3 comment.)
        """
        sp = (1 - alpha) * self.startprob + alpha * other.startprob
        tr = (1 - alpha) * self.transmat  + alpha * other.transmat
        ep = (1 - alpha) * self.emitprob  + alpha * other.emitprob
        sp = sp / sp.sum()
        tr = tr / tr.sum(axis=1, keepdims=True)
        ep = ep / ep.sum(axis=1, keepdims=True)
        return HMMParams(sp, tr, ep, self.K)


# ─────────────────────────────────────────────────────────────────────
# 1. Data loading
# ─────────────────────────────────────────────────────────────────────

def load_sessions(path: str,
                  min_sessions: int = MIN_SESSIONS
                  ) -> Dict[str, List[np.ndarray]]:
    """
    Load governance sessions from a JSONL file.

    Each line is a session with field ``steps``, each step having
    ``action.activity_type`` ∈ {ADD, UPDATE, DELETE}.

    Returns a dict {user_id: [seq1, seq2, ...]}, retaining only
    users with at least *min_sessions* valid sequences.
    """
    raw: Dict[str, List[np.ndarray]] = defaultdict(list)
    n_total = 0
    with open(path, encoding="utf-8") as fh:
        for lineno, line in enumerate(fh, 1):
            line = line.strip()
            if not line:
                continue
            try:
                s = json.loads(line)
            except json.JSONDecodeError as exc:
                print(f"[WARN] line {lineno} skipped ({exc})", file=sys.stderr)
                continue
            steps = sorted(s.get("steps", []), key=lambda x: x["step"])
            enc   = [ACTION_MAP.get(st["action"]["activity_type"], -1)
                     for st in steps]
            enc   = [e for e in enc if e >= 0]
            if len(enc) >= 2:
                uid = s.get("user_id", "unknown")
                raw[uid].append(np.array(enc, dtype=np.int32))
                n_total += 1

    users = {uid: seqs
             for uid, seqs in raw.items()
             if len(seqs) >= min_sessions}
    n_users = len(users)
    n_obs   = sum(sum(len(s) for s in seqs) for seqs in users.values())
    print(f"[load]  {n_total} total sessions, "
          f"{n_users} users (≥{min_sessions} sessions), "
          f"{n_obs} observations retained.")
    return users


def train_test_split(seqs: List[np.ndarray],
                     ratio: float = TEST_RATIO,
                     seed: int = 42
                     ) -> Tuple[List[np.ndarray], List[np.ndarray]]:
    """Stratified 1-sequence-at-a-time split."""
    rng   = np.random.default_rng(seed)
    idx   = rng.permutation(len(seqs))
    n_te  = max(1, int(len(seqs) * ratio))
    train = [seqs[i] for i in idx[n_te:]]
    test  = [seqs[i] for i in idx[:n_te]]
    return train, test


# ─────────────────────────────────────────────────────────────────────
# 2. Core HMM functions (paper-exact)
# ─────────────────────────────────────────────────────────────────────

def forward_backward(params: HMMParams,
                     seq:    np.ndarray
                     ) -> Tuple[np.ndarray, np.ndarray, float]:
    """
    Forward-backward algorithm in log-space.

    Parameters
    ----------
    params : HMMParams — θ = (π, A, B)
    seq    : (T,) int array — observed sequence

    Returns
    -------
    gamma   : (T, K)   — P(q_t=i | O, θ)
    xi_sum  : (K, K)   — Σ_{t=1}^{T-1} P(q_t=i, q_{t+1}=j | O, θ)
    log_prob: float    — log P(O | θ)
    """
    T = len(seq)
    K = params.K
    log_sp, log_tr, log_ep = params.to_log()

    # ── Forward (§4 of paper, eq. α) ──────────────────────────────
    log_alpha = np.empty((T, K))
    log_alpha[0] = log_sp + log_ep[:, seq[0]]
    for t in range(1, T):
        # log Σ_i α_{t-1}(i) A_{ij} + log B_{j,o_t}
        log_alpha[t] = (np.logaddexp.reduce(
                            log_alpha[t-1, :, None] + log_tr, axis=0)
                        + log_ep[:, seq[t]])

    log_prob = np.logaddexp.reduce(log_alpha[-1])

    # ── Backward (§4 of paper, eq. β) ─────────────────────────────
    log_beta = np.zeros((T, K))
    for t in range(T - 2, -1, -1):
        # log Σ_j A_{ij} B_{j,o_{t+1}} β_{t+1}(j)
        log_beta[t] = np.logaddexp.reduce(
            log_tr + log_ep[:, seq[t+1]] + log_beta[t+1], axis=1)

    # ── Posteriors γ and Σ_t ξ_t (§4.3 of paper) ─────────────────
    log_gamma = log_alpha + log_beta - log_prob
    gamma     = np.exp(np.clip(log_gamma, -700, 0))

    # ξ_{ij} summed over t=1..T-1  (eq. Ξ in paper, explicit bounds)
    # log_xi[t, i, j] = log_α_t(i) + log A_{ij}
    #                 + log B_{j,o_{t+1}} + log_β_{t+1}(j) - log P(O)
    if T > 1:
        log_emit_next = log_ep[:, seq[1:]].T    # (T-1, K)
        log_xi = (log_alpha[:-1, :, None]        # (T-1, K, 1)
                  + log_tr[None, :, :]           # (1,   K, K)
                  + log_emit_next[:, None, :]    # (T-1, 1, K)
                  + log_beta[1:, None, :]        # (T-1, 1, K)
                  - log_prob)
        xi_sum = np.exp(np.clip(log_xi, -700, 0)).sum(axis=0)  # (K, K)
    else:
        xi_sum = np.zeros((K, K))

    return gamma, xi_sum, float(log_prob)


def compute_local_stats(params: HMMParams,
                        seqs:   List[np.ndarray]
                        ) -> Tuple[np.ndarray, np.ndarray, np.ndarray,
                                   float, int]:
    """
    E-step for one client: compute sufficient statistics over all local sequences.

    Returns
    -------
    Gamma    : (K,)     — Σ_s γ^(s)_1
    Xi       : (K, K)   — Σ_s Σ_{t=1}^{T_s-1} ξ^(s)_t
    Phi      : (K, |X|) — Σ_s Σ_t γ^(s)_t(i) 1[o_t=o]
    total_ll : float    — Σ_s log P(O_s | θ)
    total_obs: int      — Σ_s T_s
    """
    K   = params.K
    Gamma = np.zeros(K)
    Xi    = np.zeros((K, K))
    Phi   = np.zeros((K, N_OBS))
    total_ll  = 0.0
    total_obs = 0

    for seq in seqs:
        gamma, xi_sum, lp = forward_backward(params, seq)
        Gamma      += gamma[0]                # initial state counts
        Xi         += xi_sum                  # transition counts (t=1..T-1)
        for o in range(N_OBS):               # emission counts
            Phi[:, o] += gamma[seq == o].sum(axis=0)
        total_ll  += lp
        total_obs += len(seq)

    return Gamma, Xi, Phi, total_ll, total_obs


def m_step(Gamma: np.ndarray,
           Xi:    np.ndarray,
           Phi:   np.ndarray,
           prior: float = 1e-6
           ) -> HMMParams:
    """
    M-step (§4.3 of paper, eq. mstep):
      π̂_i  = Γ_i / Σ_j Γ_j
      Â_{ij} = Ξ_{ij} / Σ_{j'} Ξ_{ij'}
      B̂_{io} = Φ_{io} / Σ_{o'} Φ_{io'}
    A small Dirichlet prior is added for numerical stability.
    """
    K = Gamma.shape[0]
    sp = Gamma + prior
    sp = sp / sp.sum()

    tr = Xi + prior
    tr = tr / tr.sum(axis=1, keepdims=True)

    ep = Phi + prior
    ep = ep / ep.sum(axis=1, keepdims=True)

    return HMMParams(sp, tr, ep, K)


def score_normalized(params: HMMParams,
                     seqs:   List[np.ndarray]) -> float:
    """Normalised log-likelihood ℓ = log P(O|θ) / n_obs."""
    if not seqs:
        return float("nan")
    total_ll  = 0.0
    total_obs = 0
    for seq in seqs:
        _, _, lp = forward_backward(params, seq)
        total_ll  += lp
        total_obs += len(seq)
    return total_ll / max(total_obs, 1)


def fit_with_hmmlearn(seqs:   List[np.ndarray],
                      K:      int,
                      n_iter: int = 100,
                      seed:   int = 42) -> HMMParams:
    """
    Fit a CategoricalHMM using hmmlearn's Baum-Welch
    (used for Local-HMM, Centralised, and FedAvg local steps).
    """
    m = CategoricalHMM(n_components=K, n_features=N_OBS,
                       n_iter=n_iter, tol=1e-4, random_state=seed)
    X       = np.concatenate([s.reshape(-1, 1) for s in seqs])
    lengths = [len(s) for s in seqs]
    m.fit(X, lengths)
    return HMMParams(m.startprob_, m.transmat_, m.emissionprob_, K)


# ─────────────────────────────────────────────────────────────────────
# 3. Algorithm [0]: Centralised (oracle upper bound)
# ─────────────────────────────────────────────────────────────────────

def run_centralised(users_train: Dict[str, List[np.ndarray]],
                    users_test:  Dict[str, List[np.ndarray]],
                    K: int = DEFAULT_K
                    ) -> Tuple[Dict, HMMParams]:
    """Single Baum-Welch on all pooled sequences."""
    print("[0] Centralised …")
    all_train = [s for seqs in users_train.values() for s in seqs]
    params    = fit_with_hmmlearn(all_train, K=K, n_iter=100, seed=0)
    results   = {}
    for uid in users_train:
        results[uid] = {
            "ll_train": score_normalized(params, users_train[uid]),
            "ll_test":  score_normalized(params, users_test[uid]),
        }
    return results, params


# ─────────────────────────────────────────────────────────────────────
# 4. Algorithm [1]: Local-HMM (baseline)
# ─────────────────────────────────────────────────────────────────────

def run_local(users_train: Dict[str, List[np.ndarray]],
              users_test:  Dict[str, List[np.ndarray]],
              K: int = DEFAULT_K
              ) -> Dict:
    """Independent Baum-Welch per user."""
    print("[1] Local-HMM …")
    results = {}
    for uid, train in users_train.items():
        p = fit_with_hmmlearn(train, K=K, n_iter=80, seed=42)
        results[uid] = {
            "ll_train": score_normalized(p, train),
            "ll_test":  score_normalized(p, users_test[uid]),
            "params":   p,
        }
    return results


# ─────────────────────────────────────────────────────────────────────
# 5. Algorithm [2]: FedAvg-HMM
# ─────────────────────────────────────────────────────────────────────

def run_fedavg(users_train: Dict[str, List[np.ndarray]],
               users_test:  Dict[str, List[np.ndarray]],
               init_params: HMMParams,
               K:        int = DEFAULT_K,
               n_rounds: int = N_ROUNDS,
               m_local:  int = M_LOCAL
               ) -> Tuple[Dict, List[float], HMMParams]:
    """
    Algorithm 2: weighted parameter averaging.

    Round r:
      1. Each client fine-tunes from θ^(r) for M local BW steps.
      2. Server: θ^(r+1) = Σ_i w_i θ^(i),  w_i = n_i / N_tot.
      3. Re-normalise rows of A, B and vector π.
    """
    print(f"[2] FedAvg-HMM  (R={n_rounds}, M={m_local}) …")
    uid_list   = list(users_train.keys())
    n_obs_dict = {uid: sum(len(s) for s in seqs)
                  for uid, seqs in users_train.items()}
    N_tot      = sum(n_obs_dict.values())

    global_p = init_params.copy()
    history: List[float] = []

    for r in range(n_rounds):
        local_params = {}
        for uid, train in users_train.items():
            # Initialise local model from global
            m = CategoricalHMM(n_components=K, n_features=N_OBS,
                               n_iter=m_local, tol=1e-6,
                               init_params='', random_state=r * 7)
            m.startprob_    = global_p.startprob.copy()
            m.transmat_     = global_p.transmat.copy()
            m.emissionprob_ = global_p.emitprob.copy()
            X = np.concatenate([s.reshape(-1, 1) for s in train])
            L = [len(s) for s in train]
            try:
                m.fit(X, L)
            except Exception:
                pass
            local_params[uid] = HMMParams(m.startprob_, m.transmat_,
                                           m.emissionprob_, K)

        # Weighted average
        sp_agg = np.zeros(K)
        tr_agg = np.zeros((K, K))
        ep_agg = np.zeros((K, N_OBS))
        for uid, lp in local_params.items():
            w = n_obs_dict[uid] / N_tot
            sp_agg += w * lp.startprob
            tr_agg += w * lp.transmat
            ep_agg += w * lp.emitprob

        # Re-normalise — technically unnecessary (convex combination of
        # row-stochastic matrices is row-stochastic: Σ_j[Σ_i w_i A^(i)_{kj}]=1)
        # but kept for numerical safety against floating-point drift.
        sp_agg /= sp_agg.sum()
        tr_agg /= tr_agg.sum(axis=1, keepdims=True)
        ep_agg /= ep_agg.sum(axis=1, keepdims=True)
        global_p = HMMParams(sp_agg, tr_agg, ep_agg, K)

        ll_mean = np.mean([score_normalized(global_p, users_train[uid])
                           for uid in uid_list])
        history.append(ll_mean)
        if (r + 1) % 5 == 0:
            print(f"    round {r+1:3d}  mean ℓ_train = {ll_mean:.4f}")

    results = {uid: {"ll_train": score_normalized(global_p, users_train[uid]),
                     "ll_test":  score_normalized(global_p, users_test[uid])}
               for uid in uid_list}
    return results, history, global_p


# ─────────────────────────────────────────────────────────────────────
# 6. Algorithm [3]: FedEM-HMM (Federated Baum-Welch)
# ─────────────────────────────────────────────────────────────────────

def run_fedem(users_train: Dict[str, List[np.ndarray]],
              users_test:  Dict[str, List[np.ndarray]],
              init_params: HMMParams,
              K:        int = DEFAULT_K,
              n_rounds: int = N_ROUNDS
              ) -> Tuple[Dict, List[float], HMMParams]:
    """
    Algorithm 1 (paper): Federated Baum-Welch.

    Round r:
      1. E-step local: each client computes raw (Γ^(i), Ξ^(i), Φ^(i))
         using the current global θ^(r).
         ── Only these aggregates are 'transmitted' to the server. ──
      2. Server aggregates:
         Γ = Σ_i Γ^(i),  Ξ = Σ_i Ξ^(i),  Φ = Σ_i Φ^(i)   ← simple sum
      3. Global M-step: θ^(r+1) = M-step(Γ, Ξ, Φ).

    Equivalence guarantee (Corollary 5.1):
      Under independence, this is algebraically identical to
      centralised Baum-Welch on the pooled dataset.

    Communication cost: K + K² + K·|X| = 32 scalars/client/round (K=4, |X|=3).
    """
    print(f"[3] FedEM-HMM   (R={n_rounds}) …")
    uid_list = list(users_train.keys())
    global_p = init_params.copy()
    history: List[float] = []

    for r in range(n_rounds):
        # ── Local E-step (distributed, parallel) ──────────────────
        agg_Gamma = np.zeros(K)
        agg_Xi    = np.zeros((K, K))
        agg_Phi   = np.zeros((K, N_OBS))

        for uid, train in users_train.items():
            Gamma_i, Xi_i, Phi_i, _, _ = compute_local_stats(global_p, train)
            # Transmit raw sufficient statistics (no normalisation)
            agg_Gamma += Gamma_i   # Γ = Σ_i Γ^(i)
            agg_Xi    += Xi_i      # Ξ = Σ_i Ξ^(i)
            agg_Phi   += Phi_i     # Φ = Σ_i Φ^(i)

        # ── Global M-step ──────────────────────────────────────────
        global_p = m_step(agg_Gamma, agg_Xi, agg_Phi)

        ll_mean = np.mean([score_normalized(global_p, users_train[uid])
                           for uid in uid_list])
        history.append(ll_mean)
        if (r + 1) % 5 == 0:
            print(f"    round {r+1:3d}  mean ℓ_train = {ll_mean:.4f}")

    results = {uid: {"ll_train": score_normalized(global_p, users_train[uid]),
                     "ll_test":  score_normalized(global_p, users_test[uid])}
               for uid in uid_list}
    return results, history, global_p


# ─────────────────────────────────────────────────────────────────────
# 7. Algorithm [4]: Personalised FedEM-HMM
# ─────────────────────────────────────────────────────────────────────

def run_personalised(users_train:    Dict[str, List[np.ndarray]],
                     users_test:     Dict[str, List[np.ndarray]],
                     global_params:  HMMParams,
                     alpha:   float = ALPHA_PERS,
                     m_p:     int   = M_PERSONAL
                     ) -> Dict:
    """
    Algorithm 3 (paper v3): Personalised FedEM.

    For each client u_i:
      1. Fine-tune from θ_global for M_p BW iterations → θ^(i)_ft.
      2. Interpolate: θ^(i)_pers = (1−α)·θ_global + α·θ^(i)_ft.
         (A convex combination of row-stochastic matrices is row-stochastic;
          renormalisation in HMMParams.interpolate is for numerical safety only.)

    α=0 → pure global model;  α=1 → pure local model.
    Heuristic: α ∝ n_i / N_tot (data-sparse clients borrow more).
    """
    print(f"[4] Personalised FedEM (α={alpha}, M_p={m_p}) …")
    results = {}
    for uid, train in users_train.items():
        # Local fine-tuning from global initialisation
        m = CategoricalHMM(n_components=global_params.K, n_features=N_OBS,
                           n_iter=m_p, tol=1e-4,
                           init_params='', random_state=42)
        m.startprob_    = global_params.startprob.copy()
        m.transmat_     = global_params.transmat.copy()
        m.emissionprob_ = global_params.emitprob.copy()
        try:
            X = np.concatenate([s.reshape(-1, 1) for s in train])
            L = [len(s) for s in train]
            m.fit(X, L)
            local_ft = HMMParams(m.startprob_, m.transmat_,
                                  m.emissionprob_, global_params.K)
        except Exception:
            local_ft = global_params.copy()

        # Convex interpolation (eq. 1 in paper)
        pers = global_params.interpolate(local_ft, alpha)
        results[uid] = {
            "ll_train": score_normalized(pers, train),
            "ll_test":  score_normalized(pers, users_test[uid]),
            "params":   pers,
        }
    return results


# ─────────────────────────────────────────────────────────────────────
# 8. Model selection (BIC / AIC)
# ─────────────────────────────────────────────────────────────────────

def model_selection(all_seqs: List[np.ndarray],
                    K_range=K_RANGE,
                    n_init: int = N_INIT
                    ) -> List[dict]:
    """
    BIC/AIC for K ∈ K_range (Table 1 in the paper).

    Correct parameter count (paper v3, §6.2):
      p(K) = (K-1) + K(K-1) + K(|X|-1)
             ─────   ──────   ──────────
             for π   for A    for B
    BIC  = -2 log L̂ + p(K) ln(N_tot)
    AIC  = -2 log L̂ + 2 p(K)
    """
    N_tot = sum(len(s) for s in all_seqs)
    log_N = math.log(N_tot)
    print(f"[model_sel]  N_tot={N_tot}, ln(N)={log_N:.4f}")
    rows = []

    for K in K_range:
        best_ll, best_p = -np.inf, None
        for seed in range(n_init):
            try:
                p = fit_with_hmmlearn(all_seqs, K=K, n_iter=100, seed=seed * 42)
                ll = sum(forward_backward(p, s)[2] for s in all_seqs)
                if ll > best_ll:
                    best_ll, best_p = ll, p
            except Exception:
                pass
        if best_p is None:
            continue
        pK  = (K - 1) + K * (K - 1) + K * (N_OBS - 1)  # π: K-1, A: K(K-1), B: K(|X|-1)
        bic = -2 * best_ll + pK * log_N
        aic = -2 * best_ll + 2 * pK
        rows.append({"K": K, "p(K)": pK, "logL": round(best_ll, 1),
                     "BIC": round(bic, 1), "AIC": round(aic, 1),
                     "params": best_p})
        print(f"  K={K}  p={pK}  logL={best_ll:.1f}  "
              f"BIC={bic:.1f}  AIC={aic:.1f}")

    return rows


# ─────────────────────────────────────────────────────────────────────
# 9. Viterbi decoding (per-session phase inference)
# ─────────────────────────────────────────────────────────────────────

STATE_LABELS = {
    0: "Intensive creation",
    1: "Creation / correction",
    2: "Deletion / cleanup",
    3: "Metadata enrichment",
}


def viterbi(params: HMMParams, seq: np.ndarray) -> np.ndarray:
    """
    Viterbi algorithm — O(K²T) — returns most probable state sequence.
    (Alg. 4 in the paper.)
    """
    T = len(seq)
    K = params.K
    log_sp, log_tr, log_ep = params.to_log()

    delta = np.empty((T, K))
    psi   = np.zeros((T, K), dtype=int)

    delta[0] = log_sp + log_ep[:, seq[0]]
    for t in range(1, T):
        scores     = delta[t-1, :, None] + log_tr       # (K, K)
        psi[t]     = np.argmax(scores, axis=0)
        delta[t]   = scores[psi[t], np.arange(K)] + log_ep[:, seq[t]]

    states = np.empty(T, dtype=int)
    states[-1] = np.argmax(delta[-1])
    for t in range(T - 2, -1, -1):
        states[t] = psi[t + 1, states[t + 1]]

    return states


def decode_all(params:    HMMParams,
               users_all: Dict[str, List[np.ndarray]]
               ) -> Dict[str, List[np.ndarray]]:
    """Viterbi-decode all sequences for all users."""
    return {uid: [viterbi(params, s) for s in seqs]
            for uid, seqs in users_all.items()}


# ─────────────────────────────────────────────────────────────────────
# 10. Excel export
# ─────────────────────────────────────────────────────────────────────

def _hcell(ws, r, c, v, sz=11):
    cl = ws.cell(r, c, v)
    cl.font      = Font(name="Arial", bold=True, color=WHITE, size=sz)
    cl.fill      = PatternFill("solid", start_color=BLUE_DARK)
    cl.border    = THIN
    cl.alignment = Alignment(horizontal="center", vertical="center")
    return cl

def _dcell(ws, r, c, v, bold=False, center=False,
           bg=None, fg="000000", size=10):
    cl = ws.cell(r, c, v)
    cl.font      = Font(name="Arial", bold=bold, size=size, color=fg)
    cl.border    = THIN
    cl.fill      = PatternFill("solid", start_color=bg if bg
                               else (ALT if r % 2 == 0 else GREY))
    cl.alignment = Alignment(horizontal="center" if center else "left",
                             vertical="center")
    return cl

def _cw(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_excel(model_sel_rows:  List[dict],
                all_results:     Dict[str, Dict],
                histories:       Dict[str, List[float]],
                uid_list:        List[str],
                users_train:     Dict[str, List[np.ndarray]],
                users_test:      Dict[str, List[np.ndarray]],
                fedem_params:    HMMParams,
                global_params:   Dict[str, Optional["HMMParams"]],
                out_path:        str) -> None:
    """
    global_params keys match all_results.
    - Single global model (Centralised, FedEM, FedAvg): global_params[algo] = HMMParams
    - Per-user models (Local, Personalised): global_params[algo] = None
      (params stored per-user in all_results[algo][uid]['params'])
    """

    wb = Workbook()
    K  = fedem_params.K
    algos = list(all_results.keys())

    # ── Sheet 1: Model selection ───────────────────────────────────
    ws1 = wb.active
    ws1.title = "1_Model_Selection"
    ws1.cell(1, 1, "Model selection — BIC / AIC").font = \
        Font(name="Arial", bold=True, size=14)
    ws1.cell(2, 1,
        f"p(K) = (K-1) + K(K-1) + K(|X|-1)  "
        f"[K-1 for π, K(K-1) for A, K(|X|-1) for B],  |X|=3,  "
        f"N_tot={sum(len(s) for seqs in users_train.values() for s in seqs)}").font = \
        Font(name="Arial", size=10, italic=True)

    for c, h in enumerate(["K", "p(K)", "log L̂", "BIC", "AIC",
                            "Selected?"], 1):
        _hcell(ws1, 4, c, h)

    best_bic_K = model_sel_rows[
        min(range(len(model_sel_rows)),
            key=lambda i: model_sel_rows[i]["BIC"])]["K"]

    for ri, row in enumerate(model_sel_rows, 5):
        is_best = row["K"] == best_bic_K
        for c, v in enumerate([row["K"], row["p(K)"], row["logL"],
                                row["BIC"], row["AIC"],
                                "✔ BIC optimal" if is_best else ""], 1):
            cl = _dcell(ws1, ri, c, v, bold=is_best, center=True)
            if is_best:
                cl.fill = PatternFill("solid", start_color="D5E8D4")
    _cw(ws1, [8, 8, 14, 12, 12, 16])

    # ── Sheet 2: Comparative results (Table 3 of paper) ───────────
    ws2 = wb.create_sheet("2_Results")
    ws2.cell(1, 1, "Comparative results — normalised test log-likelihood").font = \
        Font(name="Arial", bold=True, size=14)
    ws2.cell(2, 1,
        "ℓ_test = log P(O|θ) / n_obs  (higher is better) · "
        "Δ = improvement over Local-HMM").font = \
        Font(name="Arial", size=10, italic=True)

    hdrs = (["user_id", "n_obs (train)", "n_obs (test)"]
            + [f"ℓ_test [{a}]" for a in algos]
            + [f"Δℓ [{a}−Local]" for a in algos if a != "Local"])
    for c, h in enumerate(hdrs, 1):
        _hcell(ws2, 4, c, h)

    for ri, uid in enumerate(uid_list, 5):
        n_tr = sum(len(s) for s in users_train[uid])
        n_te = sum(len(s) for s in users_test[uid])
        ll_l = all_results.get("Local", {}).get(uid, {}).get("ll_test", float("nan"))

        row = [uid[:24], n_tr, n_te]
        for algo in algos:
            v = all_results.get(algo, {}).get(uid, {}).get("ll_test", float("nan"))
            row.append(round(v, 4) if not math.isnan(v) else "–")
        for algo in algos:
            if algo == "Local":
                continue
            v = all_results.get(algo, {}).get(uid, {}).get("ll_test", float("nan"))
            d = v - ll_l if not math.isnan(v) else float("nan")
            row.append(round(d, 4) if not math.isnan(d) else "–")

        for c, val in enumerate(row, 1):
            _dcell(ws2, ri, c, val, center=(c > 1))

        # Colour delta columns
        base = len(algos) + 4
        for j, algo in enumerate([a for a in algos if a != "Local"]):
            v = all_results.get(algo, {}).get(uid, {}).get("ll_test", float("nan"))
            d = v - ll_l if not math.isnan(v) else float("nan")
            if not math.isnan(d):
                bg = "C6EFCE" if d > 0 else "FFC7CE"
                fg = "375623" if d > 0 else "9C0006"
                _dcell(ws2, ri, base + j, round(d, 4),
                       center=True, bg=bg, fg=fg)

    _cw(ws2, [28, 14, 12] + [16] * len(algos) + [16] * (len(algos) - 1))
    ws2.freeze_panes = "A5"

    # ── Sheet 3: Summary statistics ───────────────────────────────
    ws3 = wb.create_sheet("3_Summary")
    ws3.cell(1, 1, "Summary — mean ℓ_test per algorithm").font = \
        Font(name="Arial", bold=True, size=14)

    for c, h in enumerate(["Algorithm", "Mean ℓ_test",
                            "Median ℓ_test", "Δ vs Local", "Data shared"], 1):
        _hcell(ws3, 3, c, h)

    shared_desc = {
        "Centralised": "All raw sequences",
        "Local":       "None",
        "FedAvg":      "Parameters θ = (π, A, B)",
        "FedEM":       "Sufficient stats (Γ, Ξ, Φ) only — 32 scalars/round",
        "Personalised":"Sufficient stats (Γ, Ξ, Φ) only — 32 scalars/round",
    }

    ll_local_vals = [all_results.get("Local", {}).get(uid, {}).get("ll_test",
                     float("nan")) for uid in uid_list]
    mean_local    = np.nanmean(ll_local_vals)

    for ri, algo in enumerate(algos, 4):
        vals  = [all_results.get(algo, {}).get(uid, {}).get("ll_test",
                 float("nan")) for uid in uid_list]
        mn    = round(np.nanmean(vals), 4)
        med   = round(np.nanmedian(vals), 4)
        delta = round(mn - mean_local, 4)
        color = ALGO_COLORS.get(algo, "888888")
        cl = _dcell(ws3, ri, 1, algo, bold=True)
        cl.fill = PatternFill("solid", start_color=color)
        cl.font = Font(name="Arial", bold=True, color=WHITE, size=11)
        _dcell(ws3, ri, 2, mn,    center=True)
        _dcell(ws3, ri, 3, med,   center=True)
        bg = "C6EFCE" if delta > 0 else "FFC7CE"
        fg = "375623" if delta > 0 else "9C0006"
        _dcell(ws3, ri, 4, delta, center=True, bg=bg, fg=fg)
        _dcell(ws3, ri, 5, shared_desc.get(algo, ""))
    _cw(ws3, [18, 14, 14, 12, 42])

    # ── Sheet 4: Convergence histories ────────────────────────────
    ws4 = wb.create_sheet("4_Convergence")
    ws4.cell(1, 1, "Convergence — mean ℓ_train per round").font = \
        Font(name="Arial", bold=True, size=14)
    _hcell(ws4, 3, 1, "Round")
    for c, algo in enumerate(histories.keys(), 2):
        _hcell(ws4, 3, c, algo)
    max_r = max(len(h) for h in histories.values())
    for r in range(max_r):
        _dcell(ws4, r + 4, 1, r + 1, center=True)
        for c, h in enumerate(histories.values(), 2):
            v = round(h[r], 5) if r < len(h) else "–"
            _dcell(ws4, r + 4, c, v, center=True)
    _cw(ws4, [8] + [16] * len(histories))

    # ── Sheet 5: FedEM global parameters ─────────────────────────
    ws5 = wb.create_sheet("5_FedEM_Params")
    ws5.cell(1, 1, f"FedEM global parameters θ^(R) — K={K}").font = \
        Font(name="Arial", bold=True, size=14)

    # π
    ws5.cell(3, 1, "Initial distribution π").font = \
        Font(name="Arial", bold=True, size=11)
    for i in range(K):
        _hcell(ws5, 4, i + 2, f"s{i}")
        _dcell(ws5, 5, i + 2, round(float(fedem_params.startprob[i]), 4),
               center=True)
    _dcell(ws5, 4, 1, "");  _dcell(ws5, 5, 1, "π", bold=True, center=True)

    # A
    ws5.cell(7, 1, "Transition matrix A").font = \
        Font(name="Arial", bold=True, size=11)
    _hcell(ws5, 8, 1, "from \\ to")
    for i in range(K):
        _hcell(ws5, 8, i + 2, f"s{i}")
        _hcell(ws5, 9 + i, 1, f"s{i}")
        for j in range(K):
            v  = float(fedem_params.transmat[i, j])
            cl = _dcell(ws5, 9 + i, j + 2, round(v, 4), center=True)
            intensity = int(v * 200)
            cl.fill   = PatternFill("solid",
                start_color=f"{255-intensity:02X}FF{255-intensity:02X}")

    # B
    st = 9 + K + 2
    ws5.cell(st, 1, "Emission matrix B").font = \
        Font(name="Arial", bold=True, size=11)
    _hcell(ws5, st + 1, 1, "state \\ action")
    for j, lab in enumerate(["ADD (0)", "UPDATE (1)", "DELETE (2)"], 2):
        _hcell(ws5, st + 1, j, lab)
    for i in range(K):
        _hcell(ws5, st + 2 + i, 1, f"s{i} — {STATE_LABELS.get(i,'')}")
        for j in range(N_OBS):
            v  = float(fedem_params.emitprob[i, j])
            cl = _dcell(ws5, st + 2 + i, j + 2, round(v, 4), center=True)
            intensity = int(v * 200)
            cl.fill   = PatternFill("solid",
                start_color=f"FF{255-intensity:02X}{255-intensity:02X}")
    _cw(ws5, [32] + [14] * K)

    # ── Sheet 6: Methodology ──────────────────────────────────────
    ws6 = wb.create_sheet("6_Methodology")
    ws6.cell(1, 1, "Algorithm parameters and theoretical guarantees").font = \
        Font(name="Arial", bold=True, size=14)
    notes = [
        ("Paper",         "Hemmati & Bennani, Federated Hidden Markov Models for "
                          "Privacy-Preserving Analysis of Governance Action Sequences, 2025"),
        ("Alphabet",      "X = {ADD=0, UPDATE=1, DELETE=2},  |X| = 3"),
        ("K (states)",    f"K = {K} (BIC-optimal on this dataset)"),
        ("FedEM rounds",  f"R = {N_ROUNDS}"),
        ("FedAvg M",      f"M = {M_LOCAL} local Baum-Welch iterations per round"),
        ("Personalised α",f"α = {ALPHA_PERS}  (0=global, 1=local)"),
        ("Fine-tuning M_p",f"M_p = {M_PERSONAL} iterations"),
        ("Train/test",    f"{int((1-TEST_RATIO)*100)}% / {int(TEST_RATIO*100)}% split"),
        ("FedEM exact",   "Corollary 5.1: Γ = Σ_i Γ^(i) (simple sum, no normalisation) "
                          "→ algebraically identical to centralised Baum-Welch"),
        ("Comm. cost",    f"K + K² + K|X| = {K + K**2 + K*N_OBS} scalars/client/round"),
        ("BIC formula",   "BIC = -2 log L̂ + p(K) ln(N_tot),  p(K) = K(K-1) + K(|X|-1)"),
        ("Privacy",       "For S_i ≥ 2 sequences, sufficient stats do not uniquely "
                          "determine any individual sequence → input privacy"),
    ]
    for ri, (label, note) in enumerate(notes, 3):
        _dcell(ws6, ri, 1, label, bold=True)
        cl = _dcell(ws6, ri, 2, note)
        cl.alignment = Alignment(wrap_text=True)
        ws6.row_dimensions[ri].height = 30
    _cw(ws6, [22, 90])


    # ── Sheet 7: Per-user HMM parameters ─────────────────────────
    ws7 = wb.create_sheet("7_User_HMM_Params")
    ws7.cell(1, 1, "Per-user HMM parameters θ = (π, A, B)").font =         Font(name="Arial", bold=True, size=14)
    ws7.cell(2, 1,
        "Global models (Centralised, FedEM, FedAvg) are identical for all users. "
        "Local and Personalised models are specific to each user.").font =         Font(name="Arial", size=9, italic=True)

    # ── Column headers ───────────────────────────────────────────
    hdrs7 = (["Algorithm", "User", "State", "State label",
               "π_i (initial)"]
             + [f"A→s{j}" for j in range(K)]
             + ["B(ADD)", "B(UPDATE)", "B(DELETE)"])
    for c, h in enumerate(hdrs7, 1):
        _hcell(ws7, 4, c, h)

    def _write_hmm_block(ws, start_row, algo_name, uid_label,
                         p: HMMParams, row_bg: str = None):
        """Write K rows for one (algo, user, model) block."""
        for i in range(p.K):
            r = start_row + i
            bg = row_bg if row_bg else (ALT if r % 2 == 0 else GREY)
            vals = ([algo_name if i == 0 else "",
                     uid_label if i == 0 else "",
                     f"s{i}",
                     STATE_LABELS.get(i, ""),
                     round(float(p.startprob[i]), 4)]
                    + [round(float(p.transmat[i, j]), 4) for j in range(p.K)]
                    + [round(float(p.emitprob[i, o]), 4)
                       for o in range(N_OBS)])
            for c, v in enumerate(vals, 1):
                cl = _dcell(ws, r, c, v, center=(c > 2), bg=bg)
                # Heat-map on emission columns
                if c > 5 + p.K:
                    intensity = int(float(v) * 180)
                    cl.fill = PatternFill("solid",
                        start_color=f"FF{255-intensity:02X}{255-intensity:02X}")
                # Heat-map on transition columns
                elif c > 5:
                    intensity = int(float(v) * 180)
                    cl.fill = PatternFill("solid",
                        start_color=f"{255-intensity:02X}FF{255-intensity:02X}")

    row7 = 5
    # Global models first (one block each, labelled "all users")
    global_algo_colors = {
        "Centralised": "EAD7F7",
        "FedEM":       "D5F0DC",
        "FedAvg":      "FDE9C9",
    }
    for algo in algos:
        p = global_params.get(algo)
        if p is None:
            continue   # per-user model — handled below
        bg = global_algo_colors.get(algo, ALT)
        _write_hmm_block(ws7, row7, algo, "global (all users)", p, bg)
        row7 += p.K + 1   # blank separator row

    # Per-user models
    per_user_algo_colors = {"Local": "FCE4E4", "Personalised": "D6E8F9"}
    for uid in uid_list:
        for algo in algos:
            if global_params.get(algo) is not None:
                continue   # already written above
            p = all_results.get(algo, {}).get(uid, {}).get("params")
            if p is None:
                continue
            bg = per_user_algo_colors.get(algo, ALT)
            _write_hmm_block(ws7, row7, algo, uid[:28], p, bg)
            row7 += p.K + 1

    col_widths7 = ([18, 30, 7, 26, 12]
                   + [10] * K + [10, 10, 10])
    _cw(ws7, col_widths7)
    ws7.freeze_panes = "A5"

    # ── Sheet 8: Symbolic sequences + Viterbi decoding ────────────
    ws8 = wb.create_sheet("8_Sequences")
    ws8.cell(1, 1,
        "Extracted sequences — symbolic form (A=ADD · U=UPDATE · D=DELETE) "
        "with Viterbi phase decoding (FedEM global model)").font =         Font(name="Arial", bold=True, size=14)
    ws8.cell(2, 1,
        "State labels: s0=Intensive creation · s1=Creation/correction · "
        "s2=Deletion/cleanup · s3=Metadata enrichment").font =         Font(name="Arial", size=9, italic=True)

    hdrs8 = ["user_id", "split", "session_idx", "length",
             "symbolic_sequence (A/U/D)",
             "viterbi_states (FedEM)",
             "dominant_state", "dominant_label"]
    for c, h in enumerate(hdrs8, 1):
        _hcell(ws8, 4, c, h)

    SPLIT_COLORS = {"train": "EBF5EB", "test": "EBF0FB"}
    row8 = 5
    for uid in uid_list:
        # Train sequences first, then test
        for split_name, split_seqs in [("train", users_train[uid]),
                                        ("test",  users_test[uid])]:
            bg_split = SPLIT_COLORS[split_name]
            for s_idx, seq in enumerate(split_seqs):
                # Symbolic string
                sym = "-".join(ACTION_SYM.get(int(o), "?") for o in seq)

                # Viterbi decoding with FedEM global model
                try:
                    states = viterbi(fedem_params, seq)
                    vit_str = "-".join(f"s{st}" for st in states)
                    # Dominant state (most frequent)
                    dom_st  = int(np.bincount(states, minlength=K).argmax())
                    dom_lbl = STATE_LABELS.get(dom_st, f"s{dom_st}")
                except Exception:
                    vit_str = "–"
                    dom_st  = -1
                    dom_lbl = "–"

                vals8 = [uid[:28], split_name, s_idx + 1, len(seq),
                         sym, vit_str, f"s{dom_st}", dom_lbl]
                for c, v in enumerate(vals8, 1):
                    _dcell(ws8, row8, c, v,
                           center=(c in (3, 4, 7)),
                           bg=bg_split)
                row8 += 1

        # Blank separator between users
        row8 += 1

    _cw(ws8, [30, 8, 10, 8, 60, 60, 12, 26])
    ws8.freeze_panes = "A5"

    wb.save(out_path)
    print(f"[OK]  Saved → {out_path}")


# ─────────────────────────────────────────────────────────────────────
# 11. Main pipeline
# ─────────────────────────────────────────────────────────────────────

def main() -> None:
    inp  = sys.argv[1] if len(sys.argv) > 1 else \
           "/mnt/user-data/uploads/collibra_sequences_4years.jsonl"
    outp = sys.argv[2] if len(sys.argv) > 2 else "fedhmm_results.xlsx"

    # ── 1. Load data ──────────────────────────────────────────────
    users_all = load_sessions(inp, min_sessions=MIN_SESSIONS)
    uid_list  = sorted(users_all, key=lambda u: -sum(len(s) for s in users_all[u]))

    users_train, users_test = {}, {}
    for uid, seqs in users_all.items():
        tr, te = train_test_split(seqs, ratio=TEST_RATIO, seed=42)
        users_train[uid] = tr
        users_test[uid]  = te

    for uid in uid_list:
        n_tr = sum(len(s) for s in users_train[uid])
        n_te = sum(len(s) for s in users_test[uid])
        print(f"  {uid[:20]}  train={len(users_train[uid])} sessions "
              f"({n_tr} obs)  test={len(users_test[uid])} sessions ({n_te} obs)")

    # ── 2. Model selection ────────────────────────────────────────
    print("\n── Model selection ──")
    all_train_flat = [s for seqs in users_train.values() for s in seqs]
    model_sel_rows = model_selection(all_train_flat, K_range=K_RANGE,
                                     n_init=N_INIT)
    best_K = model_sel_rows[min(range(len(model_sel_rows)),
                                key=lambda i: model_sel_rows[i]["BIC"])]["K"]
    print(f"  → K optimal (BIC) = {best_K}")

    # ── 3. Shared initialisation (from centralised BW) ───────────
    print("\n── Shared initialisation ──")
    init_p = fit_with_hmmlearn(all_train_flat, K=best_K, n_iter=50, seed=0)

    # ── 4. Run all algorithms ─────────────────────────────────────
    print("\n── Training ──")
    res_centralised, central_params = run_centralised(
        users_train, users_test, K=best_K)
    res_local   = run_local(users_train, users_test, K=best_K)
    res_fedavg, hist_fedavg, fedavg_p = run_fedavg(
        users_train, users_test, init_p,
        K=best_K, n_rounds=N_ROUNDS, m_local=M_LOCAL)
    res_fedem,  hist_fedem,  fedem_p = run_fedem(
        users_train, users_test, init_p,
        K=best_K, n_rounds=N_ROUNDS)
    res_pers = run_personalised(
        users_train, users_test, fedem_p,
        alpha=ALPHA_PERS, m_p=M_PERSONAL)

    all_results = {
        "Centralised": res_centralised,
        "Local":       res_local,
        "FedAvg":      res_fedavg,
        "FedEM":       res_fedem,
        "Personalised":res_pers,
    }
    histories = {"FedAvg": hist_fedavg, "FedEM": hist_fedem}

    # ── 5. Console summary ────────────────────────────────────────
    print("\n── Results (mean ℓ_test) ──")
    for algo, res in all_results.items():
        vals = [v.get("ll_test", float("nan")) for v in res.values()
                if not math.isnan(v.get("ll_test", float("nan")))]
        print(f"  {algo:18s}  {np.nanmean(vals):+.4f}")

    # ── 6. Export ─────────────────────────────────────────────────
    print("\n── Export Excel ──")
    # Collect per-algo model parameters
    global_params_map = {
        "Centralised": central_params,
        "Local":        None,   # per-user, in res_local[uid]['params']
        "FedAvg":       fedavg_p,
        "FedEM":        fedem_p,
        "Personalised": None,   # per-user, in res_pers[uid]['params']
    }
    build_excel(model_sel_rows, all_results, histories, uid_list,
                users_train, users_test, fedem_p,
                global_params_map, outp)


if __name__ == "__main__":
    main()
